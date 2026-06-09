"""Generate docs/Data_Sources_and_Gaps.xlsx — download links + data inventory + gaps."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.hyperlink import Hyperlink

NAVY = '0A3D62'
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill('solid', fgColor=NAVY)
BASE_FONT = Font(name='Arial', size=10)
LINK_FONT = Font(name='Arial', size=10, color='1155CC', underline='single')
WRAP_TOP = Alignment(wrap_text=True, vertical='top')
THIN = Side(style='thin', color='D8D2C4')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STATUS_FILL = {
    'Live': 'CDEBD0',
    'Statutory': 'DDE7F0',
    'Illustrative': 'FCEFC7',
    'Needed': 'FBE3DF',
    'Pending': 'ECEAE3',
}

# layer, governs/used-for, agency, dataset id, status, format/how, URL
LINKS = [
    # --- In use now (real data) ---
    ['Sidewalks (planimetric)', 'Where vending is physically allowed — the green areas', 'DCP', 'vfx9-tbb6', 'Live',
     'NYC Open Data — use the Export button (GeoJSON or Shapefile)', 'https://data.cityofnewyork.us/d/vfx9-tbb6'],
    ['Zoning Districts (C4/C5/C6)', 'Commercial-zone exclusion for General Vendors', 'DCP', 'nyzd', 'Live',
     'DCP "BYTES of the Big Apple" — Shapefile (nyzd)', 'https://www.nyc.gov/content/planning/pages/resources/datasets/gis-zoning-features'],
    ['Parks Properties', 'Out of scope (parks use a separate DPR permit system)', 'DPR', 'enfh-gkve', 'Live',
     'NYC Open Data — Export GeoJSON', 'https://data.cityofnewyork.us/d/enfh-gkve'],
    ['Subway Entrances & Exits', '10 ft no-vend buffer (both vendor types)', 'MTA', 'i9wp-a4ja', 'Live',
     'NY State Open Data — Export CSV/GeoJSON', 'https://data.ny.gov/d/i9wp-a4ja'],
    ['Fire Hydrants', '10 ft no-vend buffer (distance pending legal confirm)', 'NYCDEP', '6pui-xhxz', 'Live',
     'NYC Open Data — Export CSV (has lat/long)', 'https://data.cityofnewyork.us/d/6pui-xhxz'],
    ['Borough Boundaries', 'Borough-specific permit geography', 'DCP', 'gthc-hcne', 'Live',
     'NYC Open Data — Export GeoJSON', 'https://data.cityofnewyork.us/d/gthc-hcne'],
    ['Basemap tiles', 'Map background', 'OpenStreetMap / OpenFreeMap', '—', 'Live',
     'Open vector tiles, no API key', 'https://openfreemap.org'],

    # --- The legal "where can vendors operate" sources (gated / agency) ---
    ['DCWP General Vendor street restrictions', 'AUTHORITATIVE permitted/prohibited GV streets (the existing DCWP map)', 'DCWP', 'DS-007', 'Needed',
     'Public ArcGIS map (linked from DSNY page); WRITTEN LICENSE required for production use', 'https://www.nyc.gov/site/dsny/what-we-do/cleaning/street-vending-enforcement.page'],
    ['DOHMH Restricted Streets (Mobile Food)', 'Food time/day street restrictions (the #1 food input)', 'DOHMH', 'DS-001', 'Needed',
     'Public map online; GIS layer with hours/days via DOHMH agency request', 'https://a816-dohbesp.nyc.gov/IndicatorPublic/mobilefoodvending'],
    ['Green Cart precincts', 'Green Cart authorized geography (police-precinct based)', 'DOHMH', 'DS-005', 'Needed',
     'Precinct list/polygons via DOHMH agency request (Local Law 9 of 2008)', 'https://www.nyc.gov/site/doh/business/food-operators/mobile-and-temporary-food-vendors.page'],

    # --- Distance-buffer / out-of-scope layers (public, not yet integrated) ---
    ['Building Footprints', '20 ft building/store-entrance buffer', 'DCP', '5zhs-2jue', 'Pending',
     'NYC Open Data — Export GeoJSON', 'https://data.cityofnewyork.us/d/5zhs-2jue'],
    ['Bicycle Routes / Lanes', 'No vending in bike lanes', 'DOT', 'mzxg-pwib', 'Pending',
     'NYC Open Data — Export GeoJSON', 'https://data.cityofnewyork.us/d/mzxg-pwib'],
    ['Bus Stop Shelters', '5 ft GV buffer', 'DOT', 't4f2-8md7', 'Pending',
     'NYC Open Data — Export GeoJSON', 'https://data.cityofnewyork.us/d/t4f2-8md7'],
    ['Pedestrian Plazas (polygon)', 'Out of scope (concession-only)', 'DOT', 'k5k6-6jex', 'Pending',
     'NYC Open Data — Export GeoJSON', 'https://data.cityofnewyork.us/d/k5k6-6jex'],
    ['Parking Regulation Locations & Signs', 'No-standing zones / metered-parking rule', 'DOT', 'nfid-uabd', 'Pending',
     'NYC Open Data — Export', 'https://data.cityofnewyork.us/d/nfid-uabd'],
    ['Street Centerline (CSCL)', 'Base join key; crosswalk & corner buffers', 'DCP', 'DS-013', 'Pending',
     'NYC Open Data — search "NYC Street Centerline (CSCL)"', 'https://opendata.cityofnewyork.us/'],
    ['Sidewalk widths', '12 ft clearance rule', 'DOT / community', 'DS-040', 'Pending',
     'Derive from the sidewalk polygons (above) or community data', 'https://data.cityofnewyork.us/d/vfx9-tbb6'],
    ['Scaffolding / sidewalk sheds', 'Obstruction advisory', 'DOB', '—', 'Pending',
     'NYC Open Data — search "DOB NOW: Build – Approved Permits" (filter Sidewalk Shed)', 'https://opendata.cityofnewyork.us/'],
    ['Newsstands', '5 ft GV buffer', 'DCWP', 'DS-—', 'Pending',
     'DCWP licensed-premises / NYC Open Data — search "Newsstand"', 'https://opendata.cityofnewyork.us/'],
    ['Sidewalk cafés', '20 ft buffer', 'DCWP', 'DS-—', 'Pending',
     'NYC Open Data — search "Sidewalk Café"', 'https://opendata.cityofnewyork.us/'],
]

NEEDED = [
    ['DS-007 — DCWP General Vendor street restrictions', 'DCWP', 'Authoritative legal streets; required to replace the DCWP map', 'Execute the written license; confirm live-service vs. static export', 'P1'],
    ['DS-001 — DOHMH Restricted Streets (GIS)', 'DOHMH', 'The #1 food input; powers the time view', 'Request GIS export with per-segment hours/days', 'P1'],
    ['DS-005 — Green Cart precincts', 'DOHMH', 'Where Green Cart vendors may operate', 'Request precinct polygons (or list to convert)', 'P1'],
    ['Buffers: building entrances, bike lanes, bus shelters, plazas, no-standing, crosswalks, bus stops', 'DCP / DOT / MTA', 'Distance-buffer prohibitions + out-of-scope masks', 'Ingest the public layers (links on the first sheet); derive crosswalks from CSCL', 'P2'],
    ['Fire-hydrant stay-away distance', 'SBS Legal / FDNY', 'Engine uses 10 ft but the code only confirms a no-contact rule', 'Confirm the controlling distance', 'P2'],
    ['C-1 — WTC zone boundary', 'SBS Legal / DCWP', 'Encode the WTC hard prohibition correctly', 'Confirm Barclay vs. Vesey St (§20-465(g)(2))', 'P1 (legal)'],
    ['C-6 — Tool vs. existing agency maps', 'SBS / DSNY', 'Decides whether this replaces / federates / links the agency maps', 'Confirm the intended relationship — gates the data architecture', 'P1 (decision)'],
]


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        cell.border = BORDER
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = 'A2'


wb = Workbook()

# ---- Sheet 1: Download Links (primary) ----
ws = wb.active
ws.title = 'Download Links'
headers = ['Layer', 'What it governs / used for', 'Agency', 'Dataset ID', 'Status', 'How to download', 'Download URL']
ws.append(headers)
style_header(ws, len(headers))
for row in LINKS:
    ws.append(row)
    r = ws.max_row
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=r, column=c)
        cell.font = BASE_FONT
        cell.alignment = WRAP_TOP
        cell.border = BORDER
    ws.cell(row=r, column=5).fill = PatternFill('solid', fgColor=STATUS_FILL.get(row[4], 'FFFFFF'))
    url = row[6]
    link = ws.cell(row=r, column=7)
    if url.startswith('http'):
        link.hyperlink = Hyperlink(ref=link.coordinate, target=url)
        link.font = LINK_FONT
for col, w in zip('ABCDEFG', [26, 40, 14, 12, 12, 40, 46]):
    ws.column_dimensions[col].width = w

# ---- Sheet 2: Needed From City ----
ws2 = wb.create_sheet('Needed From City')
h2 = ['Dataset / item', 'Agency', 'Why we need it', 'Ask / blocker', 'Priority']
ws2.append(h2)
style_header(ws2, len(h2))
for row in NEEDED:
    ws2.append(row)
    r = ws2.max_row
    for c in range(1, len(h2) + 1):
        cell = ws2.cell(row=r, column=c)
        cell.font = BASE_FONT
        cell.alignment = WRAP_TOP
        cell.border = BORDER
for col, w in zip('ABCDE', [40, 18, 38, 42, 14]):
    ws2.column_dimensions[col].width = w

# ---- Sheet 3: Status Key ----
ws3 = wb.create_sheet('Status Key')
ws3.append(['Status', 'Meaning'])
style_header(ws3, 2)
for k, v in [
    ['Live', 'Real City data is downloaded, processed, and wired into the app.'],
    ['Statutory', 'Zone geometry encoded from the Admin Code text (Midtown Core, Flushing, Dyker Heights) — no file needed.'],
    ['Illustrative', 'Sample geometry in the app today; swap for the real download when obtained.'],
    ['Needed', 'Gated / agency data that governs where vendors operate — request or license required.'],
    ['Pending', 'Public layer, not yet integrated; download link provided.'],
]:
    ws3.append([k, v])
    r = ws3.max_row
    for c in (1, 2):
        ws3.cell(row=r, column=c).font = BASE_FONT
        ws3.cell(row=r, column=c).alignment = WRAP_TOP
        ws3.cell(row=r, column=c).border = BORDER
    ws3.cell(row=r, column=1).fill = PatternFill('solid', fgColor=STATUS_FILL.get(k, 'FFFFFF'))
ws3.column_dimensions['A'].width = 16
ws3.column_dimensions['B'].width = 86

wb.save('docs/Data_Sources_and_Gaps.xlsx')
print('wrote docs/Data_Sources_and_Gaps.xlsx with', len(LINKS), 'download links')
